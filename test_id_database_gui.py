import sys
import os
import sqlite3
import openpyxl
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem, QMessageBox, QFileDialog
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import QDialog, QVBoxLayout
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QComboBox
from PyQt5.QtWidgets import QDialog, QLabel, QLineEdit, QFormLayout
from PyQt5.QtWidgets import QFileDialog


# Database file path
DATABASE_FILE = '../../database/morphospecies_database.db'

# Function to create the database and tables (SQLite)
def create_sqlite_database():
    conn = sqlite3.connect(DATABASE_FILE)
    cursor = conn.cursor()

    # Create a table for the insect orders
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS insect_orders (
        id INTEGER PRIMARY KEY,
        order_name TEXT NOT NULL
    )
    """)

    # Insert some example insect orders
    orders_data = [
        (1, "Coleoptera"),
        (2, "Lepidoptera"),
        (3, "Diptera"),
        (4, "Hymenoptera"),
        (5, "Orthoptera")
    ]
    cursor.executemany("INSERT INTO insect_orders (id, order_name) VALUES (?, ?)", orders_data)

    # Create a table for the morphospecies
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS morphospecies (
        id INTEGER PRIMARY KEY,
        species_name TEXT NOT NULL,
        characters TEXT NOT NULL,
        image_path TEXT NOT NULL,
        video_file TEXT NOT NULL,
        order_id INTEGER NOT NULL,
        FOREIGN KEY (order_id) REFERENCES insect_orders(id)
    )
    """)

    # Insert some example morphospecies data
    species_data = [
        (1, "Tiger Beetle", "Fast and colorful", r"D:\Dílna\Kutění\Python\ICCS\output\visitor\LavSto2_20220524_10_14_26970_3_2_233,350_873,990.jpg;D:\Dílna\Kutění\Python\ICCS\output\visitor\LavSto2_20220524_10_14_26970_3_2_233,350_873,990.jpg;D:\Dílna\Kutění\Python\ICCS\output\visitor\LavSto2_20220524_10_14_26970_3_2_233,350_873,990.jpg", "path/to/video1", 1),
        (2, "Swallowtail Butterfly", "Large wingspan", r"D:\Dílna\Kutění\Python\ICCS\output\LavSto2_20220524_09_59_25695_2_2_211,360_851,1000.jpg", "path/to/video2", 2),
        (3, "Fruit Fly", "Tiny and agile", r"D:\Dílna\Kutění\Python\ICCS\output\visitor\LavSto2_20220524_10_14_26970_3_2_233,350_873,990.jpg;D:\Dílna\Kutění\Python\ICCS\output\visitor\LavSto2_20220524_10_14_26970_3_2_233,350_873,990.jpg;D:\Dílna\Kutění\Python\ICCS\output\visitor\LavSto2_20220524_10_14_26970_3_2_233,350_873,990.jpg", "path/to/video3", 3),
        (4, "Honey Bee", "Social insects", r"D:\Dílna\Kutění\Python\ICCS\output\visitor\LavSto2_20220524_10_14_26970_3_2_233,350_873,990.jpg", "path/to/video4", 4),
        (5, "Grasshopper", "Strong hind legs", r"D:\Dílna\Kutění\Python\ICCS\output\visitor\LavSto2_20220524_10_14_26970_3_2_233,350_873,990.jpg", "path/to/video5", 5)
    ]
    cursor.executemany("""
    INSERT INTO morphospecies (id, species_name, characters, image_path, video_file, order_id)
    VALUES (?, ?, ?, ?, ?, ?)
    """, species_data)

    # Save changes and close the connection
    conn.commit()
    conn.close()

# Function to create the database and tables (Excel)
def create_excel_database():
    workbook = openpyxl.Workbook()
    orders_sheet = workbook.create_sheet(title='insect_orders')
    orders_sheet.append(['Order ID', 'Order Name'])
    orders_data = [
        (1, 'Coleoptera'),
        (2, 'Lepidoptera'),
        (3, 'Diptera'),
        (4, 'Hymenoptera'),
        (5, 'Orthoptera')
    ]
    for order in orders_data:
        orders_sheet.append(order)

    species_sheet = workbook.create_sheet(title='morphospecies')
    species_sheet.append(['Species ID', 'Species Name', 'Characters', 'Image Path', 'Video File', 'Order ID'])
    species_data = [
        (1, 'Tiger Beetle', 'Fast and colorful', r"D:\Dílna\Kutění\Python\ICCS\output\visitor\LavSto2_20220524_10_14_26970_3_2_233,350_873,990.jpg;D:\Dílna\Kutění\Python\ICCS\output\visitor\LavSto2_20220524_10_14_26970_3_2_233,350_873,990.jpg;D:\Dílna\Kutění\Python\ICCS\output\visitor\LavSto2_20220524_10_14_26970_3_2_233,350_873,990.jpg", 'path/to/video1', 1),
        (2, 'Swallowtail Butterfly', 'Large wingspan', r"D:\Dílna\Kutění\Python\ICCS\output\visitor\LavSto2_20220524_10_14_26970_3_2_233,350_873,990.jpg", 'path/to/video2', 2),
        (3, 'Fruit Fly', 'Tiny and agile', r"D:\Dílna\Kutění\Python\ICCS\output\visitor\LavSto2_20220524_10_14_26970_3_2_233,350_873,990.jpg;D:\Dílna\Kutění\Python\ICCS\output\visitor\LavSto2_20220524_10_14_26970_3_2_233,350_873,990.jpg", 'path/to/video3', 3),
        (4, 'Honey Bee', 'Social insects', r"D:\Dílna\Kutění\Python\ICCS\output\visitor\LavSto2_20220524_10_14_26970_3_2_233,350_873,990.jpg", 'path/to/video4', 4),
        (5, 'Grasshopper', 'Strong hind legs', r"D:\Dílna\Kutění\Python\ICCS\output\visitor\LavSto2_20220524_10_14_26970_3_2_233,350_873,990.jpg", 'path/to/video5', 5)
    ]
    for species in species_data:
        species_sheet.append(species)

    workbook.save('morphospecies_database.xlsx')

# Function to load data from SQLite or Excel to the GUI
def load_data(selected_order):
    try:
        data = []

        if os.path.exists(DATABASE_FILE):
            conn = sqlite3.connect(DATABASE_FILE)
            cursor = conn.cursor()

            # Get the order ID for the selected insect order
            if selected_order == "All Orders":
                order_id = None
            else:
                cursor.execute("SELECT id FROM insect_orders WHERE order_name=?", (selected_order,))
                order_id = cursor.fetchone()
                if order_id:
                    order_id = order_id[0]

            # Fetch data based on the selected order
            if order_id is None:
                cursor.execute("SELECT * FROM morphospecies")
            else:
                cursor.execute("SELECT * FROM morphospecies WHERE order_id=?", (order_id,))

            data = [list(row) for row in cursor.fetchall()]
            conn.close()
        else:
            workbook = openpyxl.load_workbook('../../database/morphospecies_database.xlsx')
            species_sheet = workbook['morphospecies']

            # Fetch data based on the selected order
            for row in species_sheet.iter_rows(min_row=2, values_only=True):
                if selected_order == "All Orders" or row[5] == selected_order:
                    data.append(row.tolist())

        return data
    except:
        return []

# Function to save data from the GUI to SQLite or Excel
def save_data(data):
    if os.path.exists(DATABASE_FILE):
        conn = sqlite3.connect(DATABASE_FILE)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM morphospecies")
        cursor.executemany("INSERT INTO morphospecies (id, species_name, characters, image_path, video_file, order_id) VALUES (?, ?, ?, ?, ?, ?)", data)
        conn.commit()
        conn.close()
    else:
        workbook = openpyxl.load_workbook('../../database/morphospecies_database.xlsx')
        species_sheet = workbook['morphospecies']
        species_sheet.delete_rows(idx=2, amount=species_sheet.max_row)
        for species in data:
            species_sheet.append(species)
        workbook.save('morphospecies_database.xlsx')

class MorphoSpeciesApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('MorphoSpecies Database')
        self.setGeometry(100, 100, 800, 500)

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)

        layout = QVBoxLayout()

        self.table_widget = QTableWidget()
        layout.addWidget(self.table_widget)

        self.load_button = QPushButton('Load Data')
        self.load_button.clicked.connect(self.load_data_to_table)  # Connect the button click event
        layout.addWidget(self.load_button)

        self.save_button = QPushButton('Save Data')
        self.save_button.clicked.connect(self.save_data_from_table)
        layout.addWidget(self.save_button)

        self.add_button = QPushButton('Add Entry')
        self.add_button.clicked.connect(self.add_new_entry)  # Connect the button click event
        layout.addWidget(self.add_button)

        self.central_widget.setLayout(layout)

        self.selected_order = "All Orders"  # Set the initial order to "All Orders"

        self.order_combobox = QComboBox()
        self.order_combobox.addItem("All Orders")
        self.order_combobox.addItems(self.get_insect_orders())
        self.order_combobox.currentIndexChanged.connect(self.update_selected_order)
        layout.addWidget(self.order_combobox)

        # Load data int othe table
        self.load_data_to_table()

        # Load data from the database to the table
        self.load_data_to_table()

    def open_file_dialog(self, row, col):
        if col == 3:  # Image column
            file_dialog = QFileDialog(self)
            file_dialog.setFileMode(QFileDialog.ExistingFiles)
            file_dialog.setNameFilter("Images (*.png *.jpg *.jpeg *.gif)")
            if file_dialog.exec_():
                selected_files = file_dialog.selectedFiles()
                image_paths = ';'.join(selected_files)

                # Check if the QTableWidgetItem exists, create a new one if it doesn't
                item = self.table_widget.item(row, col)
                if item is None:
                    item = QTableWidgetItem(image_paths)
                    self.table_widget.setItem(row, col, item)
                else:
                    item.setText(image_paths)

                # Update the data with the new image paths
                data = load_data("All Orders")
                data[row][col] = image_paths
                save_data(data)
                self.load_data_to_table()

    def add_new_entry(self):
        dialog = QDialog(self)
        dialog.setWindowTitle('Add New Entry')

        form_layout = QFormLayout(dialog)

        species_name_input = QLineEdit()
        form_layout.addRow('Species Name:', species_name_input)

        characters_input = QLineEdit()
        form_layout.addRow('Characters:', characters_input)

        image_path_input = QLineEdit()
        form_layout.addRow('Image Path:', image_path_input)

        video_file_input = QLineEdit()
        form_layout.addRow('Video File:', video_file_input)

        order_id_input = QLineEdit()
        form_layout.addRow('Order ID:', order_id_input)

        add_button = QPushButton('Add')
        add_button.clicked.connect(
            lambda: self.add_entry_to_table(dialog, species_name_input.text(), characters_input.text(),
                                            image_path_input.text(), video_file_input.text(), order_id_input.text()))
        form_layout.addRow('', add_button)

        dialog.setLayout(form_layout)
        dialog.exec_()

    def add_entry_to_table(self, dialog, species_name, characters, image_path, video_file, order_id):
        try:
            # Get the current row count of the table
            current_row = self.table_widget.rowCount()

            # Create the new entry as a list
            new_entry = [str(current_row + 1), species_name, characters, image_path, video_file, order_id]

            # Add the new entry to the table
            self.table_widget.setRowCount(current_row + 1)
            for col, value in enumerate(new_entry):
                item = QTableWidgetItem(str(value))
                if col == 3:  # Image column
                    image_paths = value.split(';')  # Split multiple image paths by ';'
                    image_container = QWidget()
                    layout = QHBoxLayout()
                    for path in image_paths:
                        if os.path.exists(path):  # Check if the image file path exists
                            image_label = QLabel()
                            pixmap = QPixmap(path)
                            pixmap = pixmap.scaledToWidth(200)  # Adjust the image width as needed
                            image_label.setPixmap(pixmap)
                            image_label.setAlignment(Qt.AlignCenter)  # Align the image in the center
                            layout.addWidget(image_label)

                            # Add a signal to each QLabel to handle image click event
                            image_label.mousePressEvent = lambda event, path=path: self.show_enlarged_image(path)

                    image_container.setLayout(layout)
                    self.table_widget.setCellWidget(current_row, col, image_container)
                else:
                    self.table_widget.setItem(current_row, col, item)

            # Close the dialog
            dialog.close()

            # Update the database with the new entry
            data = load_data("All Orders")  # Load the existing data from the table
            max_id = max(int(entry[0]) for entry in data)  # Find the maximum ID in the existing data

            # Update the ID of the new entry to be one more than the maximum ID
            new_entry[0] = str(max_id + 1)

            # Append the new entry to the data list
            data.append(tuple(new_entry))

            # Save the updated data
            save_data(data)

        except Exception as e:
            QMessageBox.critical(self, 'Error', f'An error occurred while adding the entry: {str(e)}')

    def show_enlarged_image(self, image_path):
        if os.path.exists(image_path):
            dialog = QDialog(self)
            dialog.setWindowTitle('Enlarged Image')
            dialog_layout = QVBoxLayout(dialog)

            image_label = QLabel(dialog)
            pixmap = QPixmap(image_path)
            pixmap = pixmap.scaled(500, 500, Qt.KeepAspectRatio)  # Adjust the image size as needed
            image_label.setPixmap(pixmap)
            image_label.setAlignment(Qt.AlignCenter)
            dialog_layout.addWidget(image_label)

            dialog.setLayout(dialog_layout)
            dialog.exec_()
        else:
            QMessageBox.warning(self, 'Image Not Found', 'The image file was not found.')

    def update_selected_order(self, index):
        self.selected_order = self.order_combobox.currentText()

    def get_insect_orders(self):
        orders = []
        if os.path.exists(DATABASE_FILE):
            conn = sqlite3.connect(DATABASE_FILE)
            cursor = conn.cursor()
            cursor.execute("SELECT order_name FROM insect_orders")
            orders = [order[0] for order in cursor.fetchall()]
            conn.close()
        else:
            workbook = openpyxl.load_workbook('../../database/morphospecies_database.xlsx')
            orders_sheet = workbook['insect_orders']
            orders = [order[0].value for order in orders_sheet['B'][1:]]
        return orders

    def load_data_to_table(self):
        print("run")

        # Update the selected_order attribute
        self.selected_order = self.order_combobox.currentText()

        print(self.selected_order)
        data = load_data(self.selected_order)

        self.table_widget.clear()
        self.table_widget.setColumnCount(6)
        self.table_widget.setHorizontalHeaderLabels(
            ['Species ID', 'Species Name', 'Characters', 'Images', 'Video File', 'Order ID'])

        self.table_widget.setRowCount(len(data))

        for row, species in enumerate(data):
            for col, value in enumerate(species):
                item = QTableWidgetItem(str(value))
                if col == 3:  # Image column
                    image_paths = value.split(';')  # Split multiple image paths by ';'
                    image_container = QWidget()
                    layout = QHBoxLayout()
                    for path in image_paths:
                        if os.path.exists(path):  # Check if the image file path exists
                            image_label = QLabel()
                            pixmap = QPixmap(path)
                            pixmap = pixmap.scaledToWidth(200)  # Adjust the image width as needed
                            image_label.setPixmap(pixmap)
                            image_label.setAlignment(Qt.AlignCenter)  # Align the image in the center
                            layout.addWidget(image_label)

                            # Add a signal to each QLabel to handle image click event
                            image_label.mousePressEvent = lambda event, path=path: self.show_enlarged_image(path)

                    # Add a context menu to the image path cell
                    image_container.setContextMenuPolicy(Qt.CustomContextMenu)
                    image_container.customContextMenuRequested.connect(
                        lambda event, row=row, col=col: self.open_file_dialog(row, col))

                    image_container.setLayout(layout)
                    self.table_widget.setCellWidget(row, col, image_container)
                else:
                    self.table_widget.setItem(row, col, item)

        self.table_widget.setColumnWidth(3, 600)
        self.table_widget.verticalHeader().setDefaultSectionSize(200)

    def save_data_from_table(self):
        data = load_data("All Orders")
        # data = []
        # for row in range(self.table_widget.rowCount()):
        #     species = []
        #     for col in range(self.table_widget.columnCount()):
        #         item = self.table_widget.item(row, col)
        #         if item is not None:
        #             species.append(item.text())
        #         else:
        #             species.append("")  # Use an empty string for empty cells
        #     data.append(tuple(species))

        save_data(data)
        QMessageBox.information(self, 'Success', 'Data has been saved.')


if __name__ == '__main__':
    # Create the SQLite database if it doesn't exist
    if not os.path.exists(DATABASE_FILE):
        create_sqlite_database()

    # Create the Excel database if it doesn't exist
    if not os.path.exists('../../database/morphospecies_database.xlsx'):
        create_excel_database()

    app = QApplication(sys.argv)
    window = MorphoSpeciesApp()
    window.show()
    sys.exit(app.exec_())