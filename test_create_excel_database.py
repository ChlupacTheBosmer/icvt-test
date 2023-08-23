import openpyxl

def create_excel_database():
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Create 'insect_orders' sheet and add data
    orders_sheet = workbook.create_sheet(title='insect_orders')
    orders_sheet.append(['Order ID', 'Order Name'])
    orders_data = [
        (1, 'Coleoptera'),
        (2, 'Lepidoptera'),
        (3, 'Diptera'),
        (4, 'Hymenoptera'),
        (5, 'Orthoptera')
    ]
    for order in orders_data:
        orders_sheet.append(order)

    # Create 'morphospecies' sheet and add data
    species_sheet = workbook.create_sheet(title='morphospecies')
    species_sheet.append(['Species ID', 'Species Name', 'Characters', 'Image Path', 'Video File', 'Order ID'])
    species_data = [
        (1, 'Tiger Beetle', 'Fast and colorful', 'path/to/image1', 'path/to/video1', 1),
        (2, 'Swallowtail Butterfly', 'Large wingspan', 'path/to/image2', 'path/to/video2', 2),
        (3, 'Fruit Fly', 'Tiny and agile', 'path/to/image3', 'path/to/video3', 3),
        (4, 'Honey Bee', 'Social insects', 'path/to/image4', 'path/to/video4', 4),
        (5, 'Grasshopper', 'Strong hind legs', 'path/to/image5', 'path/to/video5', 5)
    ]
    for species in species_data:
        species_sheet.append(species)

    # Save the workbook to a file
    workbook.save('morphospecies_database.xlsx')

# Call the function to create the Excel database
create_excel_database()